from itertools import islice
import pandas as pd
from .command import section, command
from .formats import FormatIn, FormatOut
from .util import tmp_file_to_writeable, tmp_file_from_readable

section('Formats')

@command
class XlsIn(FormatIn):
  r'''
  xls-in - Read XLS Spreadsheet.
  alias: -xls

  --sheet-name=NAME  |  Sheet name
  --header, -h       |  Generate header.  Default: True.

  :suffix=.xls

  Examples:

$ psv in a.xlsx // -xls // csv-
  '''
  def format_in(self, readable, _env):
    # pylint: disable-next=import-outside-toplevel
    from openpyxl import load_workbook

    def read_workbook(filename):
      return load_workbook(filename=filename)

    workbook = tmp_file_from_readable(readable, '.xlsx', read_workbook)
    sheet_id = self.opt('sheet-name', 0)
    worksheet = workbook.worksheets[sheet_id]
    if self.opt(('header', 'h'), True):
      data = worksheet.values
      cols = next(data)[1:]
      data = list(data)
      idx = [r[0] for r in data]
      data = (islice(r, 1, None) for r in data)
      out = pd.DataFrame(data, index=idx, columns=cols)
    else:
      out = pd.DataFrame(worksheet.values)
    return out

  def default_encoding(self):
    return None

@command
class XlsOut(FormatOut):
  '''
  xls-out - Generate XLS Spreadsheet.
  alias: xls-,xls

  --sheet-name=NAME  |  Sheet name
  --header, -h       |  Generate header.  Default: True.

  :suffix=.xls

  Examples:

$ psv in a.csv // xls // o a.xlsx
$ file a.xlsx

  '''
  def format_out(self, inp, _env, writeable):
    # pylint: disable-next=import-outside-toplevel
    from openpyxl import Workbook
    # pylint: disable-next=import-outside-toplevel
    from openpyxl.utils.dataframe import dataframe_to_rows
    header = bool(self.opt('header', True))
    index = bool(self.opt('index', False))
    workbook = Workbook()
    worksheet = workbook.active

    def save_workbook(tmp_file):
      workbook.save(tmp_file)

    if isinstance(inp, pd.DataFrame):
      for row in dataframe_to_rows(inp, index=index, header=header):
        worksheet.append(row)
      tmp_file_to_writeable(writeable, '.xls', save_workbook)
    else:
      raise Exception("xls-out: cannot format {type(inp)}")

  def default_encoding(self):
    return None